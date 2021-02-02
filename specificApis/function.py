import hashlib
import json
import time
import xlwt
import openpyxl
import decimal
from specificApis import models
from django.http import HttpResponse
from datetime import datetime, timedelta, timezone, date


class configer():
    conf = models.Config.objects.get(name='default')
    startTime = str(conf.startTime)
    endTime = str(conf.endTime)
    requiredPoints = conf.requiredPoints
    pointsPerHour = conf.pointsPerHour


class MyEncoder(json.JSONEncoder):

    def default(self, obj):
        if isinstance(obj, datetime):
            return obj.strftime('%Y-%m-%d %H:%M:%S')
        elif isinstance(obj, date):
            return obj.strftime('%Y-%m-%d')
        elif isinstance(obj, decimal.Decimal):
            return float(obj)
        else:
            return json.JSONEncoder.default(self, obj)


def retJson(obj=None, mycls=json.JSONEncoder, **kwargs):
    return HttpResponse(json.dumps(kwargs if obj is None else obj, cls=mycls), content_type='application/json')


def hash(password):
    salt = '@'
    md5 = hashlib.md5()
    md5.update((password+salt).encode('utf-8'))
    return md5.hexdigest()


def UTC0_2_UTC8(utc0Time):
    # 数据库UTC时间 改成 UTC+8
    utc8Time = utc0Time.astimezone(timezone(timedelta(hours=8)))
    utc8Time = utc8Time.replace(tzinfo=None)

    return utc8Time


def check_Session(request):
    username = request.session.get('username')
    return True if username else False


def check_adminSession(request):
    username = request.session.get('username_admin')
    return True if username else False


def check_gradeAdminSession(request):
    username = request.session.get('username_grade')
    return True if username else False


def checkExist_student(studentId):
    flag = models.Student.objects.filter(studentId=studentId)
    return True if flag else False


def checkExist_admin(username):
    flag = models.Administrator.objects.filter(username=username)
    return True if flag else False


def calDurationTime(startTime, endTime):
    confi = configer()
    latestTime = datetime.strptime(confi.endTime, '%H:%M:%S')
    endTime_ = datetime.strptime(
        endTime.strftime('%H:%M:%S'), '%H:%M:%S')
    if endTime_ > latestTime:
        d = (endTime_ - latestTime).seconds
    else:
        d = 0

    duration = (endTime - startTime).seconds - d

    points = (duration / 3600) * confi.pointsPerHour
    points = round(points, 2)
    return duration, points


def check_UserPass(username, password):
    user_obj = models.User.objects.filter(username=username, password=password)
    ret = True if user_obj else False
    return ret


def check_adminPass(username, password):
    admin_obj = models.Administrator.objects.filter(
        username=username, password=password)
    ret = True if admin_obj else False
    return ret


def check_gradeAdminPass(username, password):
    admin_obj = models.GradeAdmin.objects.filter(
        username=username, password=password)
    ret = True if admin_obj else False
    return ret


def getStudentData(studentId):
    student = models.Student.objects.filter(studentId=studentId).values()[0]
    points = float(student['initPoints'])  # 总积分
    stuDatas = models.StudentData.objects.filter(
        studentId__studentId=studentId).values()

    durations = 0  # 总时长s
    for data in stuDatas:
        points += float(data['points'])
        durations += int(data['duration'])
    student['number'] = len(stuDatas)  # 总次数
    student['durations'] = durations
    student['averTime'] = round(
        durations / len(stuDatas) if len(stuDatas) else 0, 2)
    student['points'] = points

    return student, stuDatas


def getClassData(classNumber):

    c = configer()
    requiredPoints = c.requiredPoints

    students = models.Student.objects.filter(
        classNumber__classNumber=classNumber).values()
    points = 0
    durations = 0
    requiredPeople = 0
    stuDatas = []
    for stud in students:
        stuId = stud['studentId']
        stuInfo = getStudentData(stuId)[0]
        if stud['state'] == 0:  # 非免自习的人
            points = stuInfo['points']
            durations += stuInfo['averTime']
            requiredPeople += 1 if points >= requiredPoints else 0
        stuDatas.append(stuInfo)

    classs = {
        'classNumber': classNumber,
        'number': len(students),  # 人数
        'requiredPeople': requiredPeople,
        'averDurations': round(durations / len(students) if len(students) else 0, 2)
    }

    return classs, stuDatas


def cron_signOut():
    try:
        datas = models.StudentData.objects.filter(state=2).values()
        dataIds = [datas[i]['id'] for i in range(len(datas))]
        for id in dataIds:
            data = models.StudentData.objects.get(id=id)
            data.state = 4
            data.save()
            data = models.StudentData.objects.filter(id=id)
            data.update(duration=0, points=0)  # 强制结束积分为0
        return 'success'
    except Exception as e:
        return str(e)


def write_excel_xls(path, value):
    '''
    {
                "id": 35,
                "studentId": "4",
                "name": "1",
                "sex": 1,
                "state": 0,
                "initPoints": 0,
                "classNumber_id": "10011801",
                "number": 15,
                "durations": 5333,
                "averTime": 355.53,
                "points": 1.5
            },
    '''
    try:
        workbook = xlwt.Workbook()  # 新建一个工作簿
        sheet = workbook.add_sheet('sheet1')  # 在工作簿中新建一个表格
        sheet.write(0, 0, '班级')
        sheet.write(0, 1, '学号')
        sheet.write(0, 2, '姓名')
        sheet.write(0, 3, '免修')
        sheet.write(0, 4, '积分')
        sheet.write(0, 5, '自习次数')
        for i in range(0, len(value)):
            state = '是' if value[i]['state'] else '否'
            sheet.write(i+1, 0, value[i]['classNumber_id'])
            sheet.write(i+1, 1, value[i]['studentId'])
            sheet.write(i+1, 2, value[i]['name'])
            sheet.write(i+1, 3, state)
            sheet.write(i+1, 4, value[i]['points'])
            sheet.write(i+1, 5, value[i]['number'])

        workbook.save(path)  # 保存工作簿
        return 'success'
    except Exception as e:
        return str(e)


def write_excel_xlsx(path, value):
    '''
    {
                "id": 35,
                "studentId": "4",
                "name": "1",
                "sex": 1,
                "state": 0,
                "initPoints": 0,
                "classNumber_id": "10011801",
                "number": 15,
                "durations": 5333,
                "averTime": 355.53,
                "points": 1.5
            },
    '''
    try:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = 'sheet1'

        sheet.cell(row=1, column=1, value='班级')
        sheet.cell(row=1, column=2, value='学号')
        sheet.cell(row=1, column=3, value='姓名')
        sheet.cell(row=1, column=4, value='免修')
        sheet.cell(row=1, column=5, value='积分')
        sheet.cell(row=1, column=6, value='自习次数')
        for i in range(0, len(value)):
            state = '是' if value[i]['state'] else '否'
            sheet.cell(row=i+2, column=1,
                       value=str(value[i]['classNumber_id']))
            sheet.cell(row=i+2, column=2, value=str(value[i]['studentId']))
            sheet.cell(row=i+2, column=3, value=str(value[i]['name']))
            sheet.cell(row=i+2, column=4, value=str(state))
            sheet.cell(row=i+2, column=5, value=str(value[i]['points']))
            sheet.cell(row=i+2, column=6, value=str(value[i]['number']))

        workbook.save(path)
        return 'success'
    except Exception as e:
        return str(e)


def read_excel_xlsx(path):
    '''
    学号	    姓名	班级	晚上课程学时    是否免修
    2020302296	张三	10012001	0	        0
    2020302297	李四   	10012002	10	        0
    2020302298	王五  	10012003	0	        1
    '''
    ret = []
    wb = openpyxl.load_workbook(path)
    sheet = wb.worksheets[0]
    for i in range(2, sheet.max_row+1):
        stuId = sheet.cell(row=i, column=1).value
        name = sheet.cell(row=i, column=2).value
        clas = sheet.cell(row=i, column=3).value
        hours = sheet.cell(row=i, column=4).value
        state = sheet.cell(row=i, column=5).value
        if stuId == '' and name == '' and clas == '':
            break
        hours = 0 if hours == '' or not hours else int(hours)
        state = 1 if state == '1' or state == 1 else 0
        ret.append({
            'studentId': stuId,
            'name': name,
            'state': state,
            'initPoints': hours,
            'class': clas
        })
    return ret
